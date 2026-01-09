"""
Test Itinerary Generator for Visit Document Processor
Creates realistic test itineraries with real DFW addresses
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import random
import os
from datetime import datetime, timedelta

# Real DFW area addresses (verified real streets/areas)
DFW_ADDRESSES = [
    # Dallas
    ("1823 Mockingbird Ln", "Dallas", "75235"),
    ("4521 Live Oak St", "Dallas", "75204"),
    ("2910 Maple Ave", "Dallas", "75201"),
    ("6734 Greenville Ave", "Dallas", "75231"),
    ("3302 Knox St", "Dallas", "75205"),
    ("1456 Canton St", "Dallas", "75201"),
    ("5612 Gaston Ave", "Dallas", "75214"),
    ("2201 Main St", "Dallas", "75201"),
    ("4015 Cedar Springs Rd", "Dallas", "75219"),
    ("7823 Forest Ln", "Dallas", "75230"),

    # Fort Worth
    ("2134 W 7th St", "Fort Worth", "76107"),
    ("4521 Camp Bowie Blvd", "Fort Worth", "76107"),
    ("1823 Magnolia Ave", "Fort Worth", "76104"),
    ("3456 University Dr", "Fort Worth", "76109"),
    ("5612 Hulen St", "Fort Worth", "76132"),
    ("2901 Crockett St", "Fort Worth", "76107"),
    ("1234 Henderson St", "Fort Worth", "76102"),
    ("4567 Bryant Irvin Rd", "Fort Worth", "76132"),
    ("3210 Cleburne Rd", "Fort Worth", "76110"),
    ("6789 Granbury Rd", "Fort Worth", "76133"),

    # Arlington
    ("1456 Cooper St", "Arlington", "76013"),
    ("2345 S Collins St", "Arlington", "76010"),
    ("4567 W Pioneer Pkwy", "Arlington", "76013"),
    ("3234 Matlock Rd", "Arlington", "76015"),
    ("5678 Green Oaks Blvd", "Arlington", "76017"),
    ("1890 N Watson Rd", "Arlington", "76006"),
    ("2456 E Lamar Blvd", "Arlington", "76006"),
    ("3789 Little Rd", "Arlington", "76016"),

    # Irving
    ("2345 N Belt Line Rd", "Irving", "75061"),
    ("4567 MacArthur Blvd", "Irving", "75063"),
    ("1234 W Airport Fwy", "Irving", "75062"),
    ("5678 N Story Rd", "Irving", "75061"),
    ("3456 Rochelle Rd", "Irving", "75062"),

    # Garland
    ("1234 W Walnut St", "Garland", "75040"),
    ("4567 N Shiloh Rd", "Garland", "75044"),
    ("2345 Broadway Blvd", "Garland", "75043"),
    ("5678 N Garland Ave", "Garland", "75040"),
    ("3456 W Buckingham Rd", "Garland", "75042"),

    # Plano
    ("2345 W 15th St", "Plano", "75075"),
    ("4567 Coit Rd", "Plano", "75024"),
    ("1234 N Central Expy", "Plano", "75074"),
    ("5678 Preston Rd", "Plano", "75024"),
    ("3456 Legacy Dr", "Plano", "75024"),

    # Mesquite
    ("1234 N Galloway Ave", "Mesquite", "75149"),
    ("4567 E Scyene Rd", "Mesquite", "75149"),
    ("2345 Military Pkwy", "Mesquite", "75150"),

    # Grand Prairie
    ("1234 S Carrier Pkwy", "Grand Prairie", "75051"),
    ("4567 W Jefferson St", "Grand Prairie", "75051"),
    ("2345 N Great Southwest Pkwy", "Grand Prairie", "75050"),

    # Carrollton
    ("1234 E Belt Line Rd", "Carrollton", "75006"),
    ("4567 N Josey Ln", "Carrollton", "75007"),
    ("2345 Old Denton Rd", "Carrollton", "75007"),

    # Denton
    ("1234 W University Dr", "Denton", "76201"),
    ("4567 N Locust St", "Denton", "76201"),
    ("2345 Teasley Ln", "Denton", "76210"),
]

# Apartment-style addresses (add apt numbers)
APARTMENT_ADDRESSES = [
    ("500 N Akard St Apt", "Dallas", "75201", (101, 2450)),
    ("1200 Main St Apt", "Dallas", "75202", (201, 1820)),
    ("2500 Victory Park Ln Unit", "Dallas", "75219", (1001, 3540)),
    ("800 W 7th St Apt", "Fort Worth", "76102", (101, 890)),
    ("1500 S University Dr Apt", "Fort Worth", "76107", (201, 456)),
    ("400 E Abram St Apt", "Arlington", "76010", (101, 324)),
    ("600 N Center St Unit", "Arlington", "76011", (101, 256)),
    ("300 Las Colinas Blvd Apt", "Irving", "75039", (501, 2840)),
]

# First names for generating probationer/officer names
FIRST_NAMES = [
    "JAMES", "JOHN", "ROBERT", "MICHAEL", "DAVID", "WILLIAM", "RICHARD", "JOSEPH",
    "THOMAS", "CHRISTOPHER", "CHARLES", "DANIEL", "MATTHEW", "ANTHONY", "MARK",
    "DONALD", "STEVEN", "PAUL", "ANDREW", "JOSHUA", "KENNETH", "KEVIN", "BRIAN",
    "GEORGE", "TIMOTHY", "RONALD", "EDWARD", "JASON", "JEFFREY", "RYAN",
    "MARY", "PATRICIA", "JENNIFER", "LINDA", "ELIZABETH", "BARBARA", "SUSAN",
    "JESSICA", "SARAH", "KAREN", "LISA", "NANCY", "BETTY", "MARGARET", "SANDRA",
    "ASHLEY", "KIMBERLY", "EMILY", "DONNA", "MICHELLE", "DOROTHY", "CAROL",
    "AMANDA", "MELISSA", "DEBORAH", "STEPHANIE", "REBECCA", "SHARON", "LAURA"
]

LAST_NAMES = [
    "SMITH", "JOHNSON", "WILLIAMS", "BROWN", "JONES", "GARCIA", "MILLER", "DAVIS",
    "RODRIGUEZ", "MARTINEZ", "HERNANDEZ", "LOPEZ", "GONZALEZ", "WILSON", "ANDERSON",
    "THOMAS", "TAYLOR", "MOORE", "JACKSON", "MARTIN", "LEE", "PEREZ", "THOMPSON",
    "WHITE", "HARRIS", "SANCHEZ", "CLARK", "RAMIREZ", "LEWIS", "ROBINSON",
    "WALKER", "YOUNG", "ALLEN", "KING", "WRIGHT", "SCOTT", "TORRES", "NGUYEN",
    "HILL", "FLORES", "GREEN", "ADAMS", "NELSON", "BAKER", "HALL", "RIVERA",
    "CAMPBELL", "MITCHELL", "CARTER", "ROBERTS", "GOMEZ", "PHILLIPS", "EVANS"
]

# Visit types with weights (probability distribution)
VISIT_TYPES = [
    ("AM Intake", 0.25),
    ("PM Intake", 0.20),
    ("FTR", 0.25),
    ("HR", 0.10),
    ("HR AM", 0.08),
    ("HR PM", 0.07),
    ("CM", 0.05),
]

# Officer pairs (these will be the assigned officers)
OFFICER_PAIRS = [
    ("JOHNSON, SARAH", "MARTINEZ, DAVID"),
    ("SMITH, MICHAEL", "GARCIA, JENNIFER"),
    ("WILLIAMS, ROBERT", "BROWN, ASHLEY"),
    ("DAVIS, THOMAS", "RODRIGUEZ, MARIA"),
    ("WILSON, JAMES", "ANDERSON, LISA"),
]


def generate_name():
    """Generate a random name in LAST, FIRST format"""
    return f"{random.choice(LAST_NAMES)}, {random.choice(FIRST_NAMES)}"


def generate_phone():
    """Generate a random DFW-area phone number"""
    area_codes = ["214", "972", "469", "817", "682", "940"]
    return f"({random.choice(area_codes)}) {random.randint(200, 999)}-{random.randint(1000, 9999)}"


def get_random_address():
    """Get a random address from the pool"""
    # 70% chance of regular address, 30% chance of apartment
    if random.random() < 0.7:
        addr, city, zip_code = random.choice(DFW_ADDRESSES)
        return addr, city, zip_code
    else:
        addr_base, city, zip_code, apt_range = random.choice(APARTMENT_ADDRESSES)
        apt_num = random.randint(apt_range[0], apt_range[1])
        return f"{addr_base} {apt_num}", city, zip_code


def get_visit_type():
    """Get a random visit type based on weighted probability"""
    rand = random.random()
    cumulative = 0
    for vtype, weight in VISIT_TYPES:
        cumulative += weight
        if rand <= cumulative:
            return vtype
    return VISIT_TYPES[0][0]  # Default to AM Intake


def generate_itinerary(num_visits, output_path, officer_pair=None):
    """
    Generate a test itinerary Excel file

    Args:
        num_visits: Number of visits to include
        output_path: Path to save the Excel file
        officer_pair: Tuple of (officer1, officer2) or None for random
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Itinerary"

    # Select officer pair
    if officer_pair is None:
        officer_pair = random.choice(OFFICER_PAIRS)

    # Style setup
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers (row 7 to match real itinerary format)
    headers = ["Loc", "Unit", "Defendant", "Officer", "Starts",
               "Cell Phone", "Address", "City", "Zip", "Comment"]

    header_row = 7
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Track used addresses to avoid duplicates within same itinerary
    used_addresses = set()
    used_names = set()

    # Generate visit data
    start_time = datetime.strptime("8:00 AM", "%I:%M %p")

    for i in range(num_visits):
        row = header_row + 1 + i

        # Loc (numerical order)
        ws.cell(row=row, column=1, value=i + 1)

        # Unit (alphabetical order starting with B)
        unit_letter = chr(ord('B') + i)
        ws.cell(row=row, column=2, value=unit_letter)

        # Defendant (unique name)
        defendant = generate_name()
        while defendant in used_names:
            defendant = generate_name()
        used_names.add(defendant)
        ws.cell(row=row, column=3, value=defendant)

        # Officer (alternating between the pair)
        officer = officer_pair[i % 2]
        ws.cell(row=row, column=4, value=officer)

        # Starts (every 30 mins from 8:00 AM)
        visit_time = start_time + timedelta(minutes=30 * i)
        ws.cell(row=row, column=5, value=visit_time.strftime("%I:%M %p"))

        # Cell Phone
        ws.cell(row=row, column=6, value=generate_phone())

        # Address (unique within itinerary)
        address, city, zip_code = get_random_address()
        attempts = 0
        while address in used_addresses and attempts < 50:
            address, city, zip_code = get_random_address()
            attempts += 1
        used_addresses.add(address)

        ws.cell(row=row, column=7, value=address)
        ws.cell(row=row, column=8, value=city)
        ws.cell(row=row, column=9, value=zip_code)

        # Comment (visit type)
        ws.cell(row=row, column=10, value=get_visit_type())

        # Apply borders to all cells in row
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin_border

    # Adjust column widths
    column_widths = [6, 6, 25, 25, 12, 16, 35, 15, 8, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Save
    wb.save(output_path)
    print(f"Created: {output_path} with {num_visits} visits")
    return output_path


def generate_updated_itinerary(original_path, num_additional, output_path):
    """
    Create an updated itinerary that includes original visits plus new ones
    Used for testing "Add Updated Itinerary" feature

    Args:
        original_path: Path to original itinerary
        num_additional: Number of additional visits to add
        output_path: Path to save updated itinerary
    """
    # Load original
    wb = openpyxl.load_workbook(original_path)
    ws = wb.active

    # Find last row with data
    header_row = 7
    last_row = header_row
    for row in range(header_row + 1, 100):
        if ws.cell(row=row, column=1).value is not None:
            last_row = row
        else:
            break

    # Get existing data for continuation
    existing_count = last_row - header_row

    # Get the officer from existing data
    existing_officer = ws.cell(row=header_row + 1, column=4).value
    officer_pair = (existing_officer, existing_officer)  # Simplified

    # Track used addresses and names
    used_addresses = set()
    used_names = set()
    for row in range(header_row + 1, last_row + 1):
        used_addresses.add(ws.cell(row=row, column=7).value)
        used_names.add(ws.cell(row=row, column=3).value)

    # Get last time
    last_time_str = ws.cell(row=last_row, column=5).value
    if last_time_str:
        try:
            last_time = datetime.strptime(last_time_str, "%I:%M %p")
        except:
            last_time = datetime.strptime("12:00 PM", "%I:%M %p")
    else:
        last_time = datetime.strptime("12:00 PM", "%I:%M %p")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add new visits
    for i in range(num_additional):
        row = last_row + 1 + i

        # Loc (continuing numerical order)
        ws.cell(row=row, column=1, value=existing_count + i + 1)

        # Unit (continuing alphabetical)
        unit_letter = chr(ord('B') + existing_count + i)
        if ord(unit_letter) > ord('Z'):
            unit_letter = f"A{chr(ord('A') + (existing_count + i - 24))}"
        ws.cell(row=row, column=2, value=unit_letter)

        # Defendant (unique)
        defendant = generate_name()
        while defendant in used_names:
            defendant = generate_name()
        used_names.add(defendant)
        ws.cell(row=row, column=3, value=defendant)

        # Officer
        ws.cell(row=row, column=4, value=existing_officer)

        # Time (continuing from last time)
        visit_time = last_time + timedelta(minutes=30 * (i + 1))
        ws.cell(row=row, column=5, value=visit_time.strftime("%I:%M %p"))

        # Phone
        ws.cell(row=row, column=6, value=generate_phone())

        # Address
        address, city, zip_code = get_random_address()
        while address in used_addresses:
            address, city, zip_code = get_random_address()
        used_addresses.add(address)

        ws.cell(row=row, column=7, value=address)
        ws.cell(row=row, column=8, value=city)
        ws.cell(row=row, column=9, value=zip_code)

        # Visit type
        ws.cell(row=row, column=10, value=get_visit_type())

        # Borders
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin_border

    wb.save(output_path)
    print(f"Created updated itinerary: {output_path} (+{num_additional} visits)")
    return output_path


def create_test_set():
    """Create a complete set of test itineraries"""
    base_path = os.path.dirname(os.path.abspath(__file__))
    itinerary_path = os.path.join(base_path, "test_itineraries")

    # Ensure directory exists
    os.makedirs(itinerary_path, exist_ok=True)

    # Create various sized itineraries
    sizes = [5, 8, 12, 18]

    created_files = []

    for size in sizes:
        # Main itinerary
        main_file = os.path.join(itinerary_path, f"test_itinerary_{size}_visits.xlsx")
        generate_itinerary(size, main_file)
        created_files.append(main_file)

        # Updated version with 2-3 additional visits
        additional = random.randint(2, 3)
        updated_file = os.path.join(itinerary_path, f"test_itinerary_{size}_updated_{additional}more.xlsx")
        generate_updated_itinerary(main_file, additional, updated_file)
        created_files.append(updated_file)

    print(f"\n{'='*60}")
    print("Test itinerary set created!")
    print(f"{'='*60}")
    print(f"Location: {itinerary_path}")
    print(f"Files created: {len(created_files)}")
    for f in created_files:
        print(f"  - {os.path.basename(f)}")

    return created_files


if __name__ == "__main__":
    create_test_set()
